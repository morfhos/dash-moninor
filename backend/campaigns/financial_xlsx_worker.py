"""
Subprocess worker: parse financial Excel (Integração Financeira Fase 2).

Handles the Tactical Media Plan format with tabs like:
  - RESUMO DE MEIOS *  → split by channel from TT rows + per-vehicle detail
  - TV ABERTA *        → media efficiency rows (MARKET/CHANNEL/PROGRAMM/days)
  - TV PAGA *          → media efficiency rows
  - RÁDIO *            → media efficiency rows
  - JORNAL             → media efficiency rows
  - DIGITAL            → media efficiency rows
  - CUSTO GERAÇÃO *    → generation costs
  - COVER              → campaign metadata

Run as: python -m campaigns.financial_xlsx_worker <path>
Output: JSON to stdout
"""

from __future__ import annotations

import json
import re
import sys
from typing import Any


def _safe_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        if isinstance(v, str):
            v = v.replace(",", ".").strip().rstrip("%")
        return float(v)
    except (ValueError, TypeError):
        return None


def _safe_int(v: Any) -> int | None:
    f = _safe_float(v)
    if f is None:
        return None
    return int(round(f))


def _norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    return (
        s.replace("á", "a").replace("à", "a").replace("â", "a").replace("ã", "a")
         .replace("é", "e").replace("ê", "e").replace("í", "i")
         .replace("ó", "o").replace("ô", "o").replace("õ", "o")
         .replace("ú", "u").replace("ç", "c")
    )


def _find_sheets(wb, pattern: str):
    """Find all sheets whose normalized name contains pattern."""
    pat = _norm(pattern)
    return [wb[s] for s in wb.sheetnames if pat in _norm(s)]


def _find_sheet(wb, candidates: list[str]):
    """Find first sheet matching any candidate (substring match)."""
    for c in candidates:
        found = _find_sheets(wb, c)
        if found:
            return found[0]
    return None


CHANNEL_ALIASES = {
    "tv aberta": "tv_aberta",
    "open tv": "tv_aberta",
    "tv paga": "paytv",
    "tv fechada": "paytv",
    "pay tv": "paytv",
    "paytv": "paytv",
    "radio": "radio",
    "radios": "radio",
    "jornal": "jornal",
    "jornais": "jornal",
    "digital": "digital",
    "ooh": "ooh",
    "midia exterior": "ooh",
    "geracao": "geracao",
    "custo geracao": "geracao",
}

MONTH_KEYWORDS = {
    "janeiro": 1, "fevereiro": 2, "marco": 3, "abril": 4, "maio": 5, "junho": 6,
    "julho": 7, "agosto": 8, "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12,
}


# ── RESUMO DE MEIOS ───────────────────────────────────────────────────────────

def parse_resumo_meios(wb) -> tuple[dict, dict, list[dict]]:
    """
    Parse all RESUMO DE MEIOS tabs.
    Returns (channel_totals, monthly_investment, efficiency_rows).

    Format: header row with PRAÇA, MEIO, VEÍCULO, <months>, TOTAL BRUTO, VALOR DESEMBOLSO, PART GERAL %
    Data rows: per-vehicle lines. TT rows (TT TV ABERTA, TT RÁDIO, etc.) are channel totals.
    """
    sheets = _find_sheets(wb, "resumo de meios")
    if not sheets:
        return {}, {}, []

    channel_totals: dict[str, dict] = {}
    monthly: dict[str, float] = {}  # month_key -> accumulated value
    eff_rows: list[dict] = []

    for ws in sheets:
        sheet_label = _norm(ws.title)

        # Find header row with PRAÇA/MEIO/VEÍCULO
        header_row = None
        for r in range(1, 15):
            row_vals = [_norm(str(ws.cell(row=r, column=c).value or "")) for c in range(1, 15)]
            if any("praca" in v or "market" in v for v in row_vals) and any("meio" in v or "veiculo" in v for v in row_vals):
                header_row = r
                break
        if header_row is None:
            continue

        # Map columns
        col_map: dict[str, int] = {}
        month_cols: list[tuple[int, int]] = []  # (col, month_num)
        for c in range(1, ws.max_column + 1):
            h = _norm(str(ws.cell(row=header_row, column=c).value or ""))
            if not h:
                continue
            if "praca" in h or "market" in h:
                col_map["praca"] = c
            elif h == "meio":
                col_map["meio"] = c
            elif "veiculo" in h:
                col_map["veiculo"] = c
            elif "total bruto" in h:
                col_map["total_bruto"] = c
            elif "desembolso" in h:
                col_map["desembolso"] = c
            elif "part" in h and "geral" in h:
                col_map["part_pct"] = c
            else:
                # Check if it's a month name
                for mk, mn in MONTH_KEYWORDS.items():
                    if mk in h:
                        month_cols.append((c, mn))
                        break

        carry_praca = ""
        carry_meio = ""

        for r in range(header_row + 1, ws.max_row + 1):
            # Read raw cell values
            praca_val = ws.cell(row=r, column=col_map.get("praca", 2)).value
            meio_val = ws.cell(row=r, column=col_map.get("meio", 3)).value
            veiculo_val = ws.cell(row=r, column=col_map.get("veiculo", 4)).value
            total_bruto = _safe_float(ws.cell(row=r, column=col_map.get("total_bruto", 8)).value)
            desembolso = _safe_float(ws.cell(row=r, column=col_map.get("desembolso", 9)).value)
            part_pct = _safe_float(ws.cell(row=r, column=col_map.get("part_pct", 10)).value)

            praca_str = str(praca_val or "").strip()
            meio_str = str(meio_val or "").strip()
            veiculo_str = str(veiculo_val or "").strip()

            # Carry-over for merged cells
            if praca_str:
                carry_praca = praca_str
            if meio_str:
                carry_meio = meio_str

            label_norm = _norm(praca_str)

            # Check if this is a TT (total) row
            if label_norm.startswith("tt ") or "total geral" in label_norm:
                channel_key = None
                for alias, key in CHANNEL_ALIASES.items():
                    if alias in label_norm:
                        channel_key = key
                        break

                if "total geral" in label_norm:
                    # Overall total
                    channel_key = "_total"

                if channel_key and (total_bruto is not None or desembolso is not None):
                    existing = channel_totals.get(channel_key, {})
                    existing["valor_bruto"] = (existing.get("valor_bruto") or 0) + (total_bruto or 0)
                    existing["valor_liquido"] = (existing.get("valor_liquido") or 0) + (desembolso or 0)
                    if part_pct is not None:
                        existing["part_pct"] = part_pct
                    channel_totals[channel_key] = existing

                    # Accumulate monthly values from TT rows for timeline
                    for col, month_num in month_cols:
                        v = _safe_float(ws.cell(row=r, column=col).value)
                        if v and v > 0 and channel_key != "_total":
                            mk = f"2026-{month_num:02d}"
                            monthly[mk] = monthly.get(mk, 0) + v

                continue

            # Regular vehicle row
            if not veiculo_str and not meio_str:
                continue
            if total_bruto is None and desembolso is None:
                continue

            eff_rows.append({
                "praca": carry_praca,
                "channel_type": _resolve_channel(carry_meio),
                "veiculo": veiculo_str or carry_meio,
                "custo_tabela": total_bruto,
                "custo_negociado": desembolso,
                "valor": desembolso,
                "ia_pct": part_pct,
            })

    monthly_list = [{"month": k, "valor": v} for k, v in sorted(monthly.items()) if v > 0]
    return channel_totals, {"monthly_investment": monthly_list} if monthly_list else {}, eff_rows


def _resolve_channel(meio: str) -> str:
    n = _norm(meio)
    for alias, key in CHANNEL_ALIASES.items():
        if alias in n:
            return key
    return "other"


# ── CUSTO GERAÇÃO tab ─────────────────────────────────────────────────────────

def parse_custo_geracao(wb) -> list[dict]:
    """Parse CUSTO GERAÇÃO tabs (envio de material costs)."""
    sheets = _find_sheets(wb, "custo geracao")
    if not sheets:
        sheets = _find_sheets(wb, "geracao")
    if not sheets:
        return []

    rows: list[dict] = []
    for ws in sheets:
        # Find header row
        header_row = None
        for r in range(1, 10):
            row_vals = [_norm(str(ws.cell(row=r, column=c).value or "")) for c in range(1, 12)]
            if any("veiculo" in v for v in row_vals) and any("custo" in v for v in row_vals):
                header_row = r
                break
        if header_row is None:
            continue

        col_map: dict[str, int] = {}
        for c in range(1, 12):
            h = _norm(str(ws.cell(row=header_row, column=c).value or ""))
            if "veiculo" in h:
                col_map["veiculo"] = c
            elif "descricao" in h or "descrição" in h:
                col_map["descricao"] = c
            elif "destino" in h:
                col_map["destino"] = c
            elif "uf" in h:
                col_map["uf"] = c
            elif "material" in h:
                col_map["material"] = c
            elif "custo unit" in h:
                col_map["custo_unit"] = c
            elif "qtd" in h:
                col_map["qtd"] = c
            elif "custo liquido" in h or "custo liq" in h:
                col_map["custo_liq"] = c

        carry_veiculo = ""
        for r in range(header_row + 1, ws.max_row + 1):
            veiculo_val = ws.cell(row=r, column=col_map.get("veiculo", 2)).value
            custo_liq = _safe_float(ws.cell(row=r, column=col_map.get("custo_liq", 9)).value)

            veiculo_str = str(veiculo_val or "").strip()
            if veiculo_str:
                carry_veiculo = veiculo_str

            # Skip TT GERAL and empty
            if "tt geral" in _norm(str(ws.cell(row=r, column=col_map.get("custo_liq", 6)).value or "")):
                continue
            if custo_liq is None or custo_liq == 0:
                continue
            if not veiculo_str and not ws.cell(row=r, column=col_map.get("destino", 4)).value:
                continue

            destino = str(ws.cell(row=r, column=col_map.get("destino", 4)).value or "").strip()
            material = str(ws.cell(row=r, column=col_map.get("material", 6)).value or "").strip()
            uf = str(ws.cell(row=r, column=col_map.get("uf", 5)).value or "").strip()
            qtd = _safe_int(ws.cell(row=r, column=col_map.get("qtd", 8)).value)

            rows.append({
                "veiculo": carry_veiculo,
                "destino": destino,
                "uf": uf,
                "material": material,
                "custo_unitario": _safe_float(ws.cell(row=r, column=col_map.get("custo_unit", 7)).value),
                "qtd": qtd or 0,
                "custo_liquido": custo_liq,
            })

    return rows


# ── MAIN ──────────────────────────────────────────────────────────────────────

def _disable_lxml() -> None:
    """Block lxml import to prevent segfault on Python 3.8 + openpyxl."""
    try:
        import importlib.abc
        import importlib.machinery
    except Exception:
        return

    class _BlockLxml(importlib.abc.MetaPathFinder, importlib.abc.Loader):
        def find_spec(self, fullname: str, path: Any, target: Any = None) -> Any:
            if fullname == "lxml" or fullname.startswith("lxml."):
                return importlib.machinery.ModuleSpec(fullname, self)
            return None

        def create_module(self, spec: Any) -> None:
            return None

        def exec_module(self, module: Any) -> None:
            raise ImportError("lxml blocked")

    sys.meta_path.insert(0, _BlockLxml())


def main(path: str) -> dict:
    _disable_lxml()
    try:
        import openpyxl
    except ImportError:
        return {"ok": False, "errors": ["openpyxl not installed"]}

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        return {"ok": False, "errors": [f"Cannot open file: {e}"]}

    sheets_found = wb.sheetnames

    # Parse RESUMO DE MEIOS → channel totals + monthly + per-vehicle efficiency rows
    channel_totals, summary_data, eff_rows = parse_resumo_meios(wb)

    # Derive financial summary from channel totals
    total_row = channel_totals.pop("_total", {})
    summary: dict[str, Any] = {}
    summary["total_valor_tabela"] = total_row.get("valor_bruto")
    summary["total_desembolso"] = total_row.get("valor_liquido")
    if summary.get("total_valor_tabela") and summary.get("total_desembolso"):
        vt = summary["total_valor_tabela"]
        vd = summary["total_desembolso"]
        summary["total_valor_negociado"] = vd
        if vt > 0:
            summary["desconto_pct"] = round((1 - vd / vt) * 100, 2)

    # Merge monthly investment
    if summary_data.get("monthly_investment"):
        summary["monthly_investment"] = summary_data["monthly_investment"]

    # Parse CUSTO GERAÇÃO
    geracao_rows = parse_custo_geracao(wb)

    # Media efficiency: combine resumo rows
    media_efficiencies = []
    for row in eff_rows:
        media_efficiencies.append(row)

    # Aggregate investment by praça (region)
    region_totals: dict[str, float] = {}
    for row in eff_rows:
        praca = (row.get("praca") or "").strip()
        if not praca:
            continue
        v = row.get("valor") or row.get("custo_negociado") or 0
        if v:
            region_totals[praca] = region_totals.get(praca, 0) + v

    # Convert to list with percentages
    grand_total = sum(region_totals.values()) or 1
    region_list = []
    for praca, valor in sorted(region_totals.items(), key=lambda x: -x[1]):
        region_list.append({
            "region_name": praca,
            "valor": round(valor, 2),
            "percentage": round((valor / grand_total) * 100, 2),
        })

    return {
        "ok": True,
        "sheets_found": sheets_found,
        "summary": summary,
        "resumo_meios": channel_totals,
        "pi_controls": [],
        "media_efficiencies": media_efficiencies,
        "custo_geracao": geracao_rows,
        "region_investments": region_list,
    }


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"ok": False, "errors": ["Usage: financial_xlsx_worker.py <path>"]}))
        sys.exit(1)
    result = main(sys.argv[1])
    print(json.dumps(result, ensure_ascii=False, default=str))
