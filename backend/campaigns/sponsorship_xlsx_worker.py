"""
Parser para planilhas de patrocínio/proposta comercial.

Suporta dois formatos:
  - "globo"  → colunas de mês (FEV, MAR, ABR...) com MARCA, PLATAFORMA, FORMATO
  - "valor"  → colunas PERIODO(mês) + INSERCOES com ACAO/TITULO, PLATAFORMA, FORMATO

Produz o mesmo JSON que xlsx_worker.py para reuso do pipeline de import.
Campos extras por linha:
  data["day_costs"]       → lista de float|None paralela a days
  data["day_impressions"] → lista de int|None paralela a days
"""
from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime
from typing import Any


# ─── Normalização ────────────────────────────────────────────────────────────

def _norm(s: Any) -> str:
    s = str(s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    return (
        s.replace("á", "a").replace("à", "a").replace("â", "a").replace("ã", "a")
         .replace("é", "e").replace("ê", "e").replace("í", "i")
         .replace("ó", "o").replace("ô", "o").replace("õ", "o")
         .replace("ú", "u").replace("ç", "c")
    )


_MONTHS: dict[str, int] = {
    "jan": 1, "janeiro": 1,
    "fev": 2, "fevereiro": 2,
    "mar": 3, "marco": 3, "março": 3,
    "abr": 4, "abril": 4,
    "mai": 5, "maio": 5,
    "jun": 6, "junho": 6,
    "jul": 7, "julho": 7,
    "ago": 8, "agosto": 8,
    "set": 9, "setembro": 9,
    "out": 10, "outubro": 10,
    "nov": 11, "novembro": 11,
    "dez": 12, "dezembro": 12,
}


def _month_from_str(s: Any) -> int | None:
    n = _norm(s)
    for name, num in _MONTHS.items():
        if n == name or n.startswith(name):
            return num
    return None


def _parse_int(v: Any) -> int | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, int):
        return int(v)
    if isinstance(v, float):
        return int(round(v))
    s = str(v).strip()
    if not s:
        return None
    s = s.replace(",", ".")
    try:
        return int(round(float(s)))
    except Exception:
        return None


def _parse_cost(v: Any) -> float | None:
    """Parseia valor monetário. 'BONIFICADO' ou '-' retorna 0.0."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().upper()
    if not s or s in {"-", "N/A", "NA", "0"}:
        return 0.0
    if "BONIFICADO" in s or "BONI" in s:
        return 0.0
    # Remove símbolos monetários
    s = re.sub(r"[R$\s]", "", s)
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


# ─── Inferência de ano ────────────────────────────────────────────────────────

def _infer_year(ws: Any, until_row: int) -> int:
    for r in range(1, max(1, min(until_row, ws.max_row)) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            m = re.search(r"\b(20\d{2})\b", str(v))
            if m:
                y = int(m.group(1))
                if 2020 <= y <= 2035:
                    return y
    return datetime.now().year


# ─── Mapeamento PLATAFORMA → (media_type, media_channel) ─────────────────────

def _plataforma_to_channel(plataforma: Any) -> tuple[str, str]:
    n = _norm(plataforma or "")
    if any(k in n for k in ("impresso", "jornal", "print", "revista", "diario")):
        return ("offline", "jornal")
    if any(k in n for k in ("radio",)):
        return ("offline", "radio")
    if any(k in n for k in ("tv", "televisao")):
        return ("offline", "tv_aberta")
    if any(k in n for k in ("youtube", "video")):
        return ("online", "youtube")
    if any(k in n for k in ("search", "busca")):
        return ("online", "search")
    if any(k in n for k in ("display", "banner", "programatica")):
        return ("online", "display")
    if any(k in n for k in ("redes sociais", "social", "instagram", "facebook", "linkedin", "twitter", "tiktok")):
        return ("online", "social")
    if any(k in n for k in ("site", "portal", "digital", "newsletter", "email", "e-mail")):
        return ("online", "display")
    return ("online", "other")


# ─── Detecção do cabeçalho da planilha ───────────────────────────────────────

_SKIP_SHEETS = {"resumo", "planilha", "sheet", "sumario", "sumário", "total"}


def _should_skip_sheet(title: str) -> bool:
    n = _norm(title)
    return any(s in n for s in _SKIP_SHEETS)


def _find_sponsorship_header(ws: Any) -> tuple[int, str, dict[str, int]] | None:
    """
    Encontra a linha de cabeçalho de uma planilha de patrocínio.
    Retorna (row_idx, format_type, col_map) ou None.
    format_type é "globo" ou "valor".
    """
    for r in range(1, min(ws.max_row, 20) + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        norm_row = [_norm(v) for v in row_vals]

        # Precisa ter "plataforma" no cabeçalho
        if "plataforma" not in norm_row:
            continue

        col_map: dict[str, int] = {}
        month_cols: dict[int, int] = {}  # month_num → col_idx

        for i, nv in enumerate(norm_row, start=1):
            if not nv:
                continue
            if nv == "plataforma":
                col_map["plataforma"] = i
            elif "marca" == nv:
                col_map["marca"] = i
            elif nv in {"acao", "ação"}:
                col_map["acao"] = i
            elif nv == "titulo" or nv == "título":
                col_map["titulo"] = i
            elif "projeto" in nv or "entrega" in nv:
                col_map["projeto"] = i
            elif "caderno" in nv or "editoria" in nv:
                col_map["caderno"] = i
            elif nv == "formato":
                col_map["formato"] = i
            elif "periodo" in nv or "período" in nv:
                col_map["periodo"] = i
            elif "insercoes" in nv or "inserções" in nv:
                if "total" not in nv:
                    col_map["insercoes"] = i
                else:
                    col_map["total_insercoes"] = i
            elif "media" in nv and "impacto" in nv:
                col_map["media_impactos"] = i
            elif ("valor" in nv or "vt" == nv) and "tabela" in nv:
                col_map["valor_tabela"] = i
            elif "negociado" in nv and "liquido" in nv:
                col_map["valor_negociado_liquido"] = i
            elif "negociado" in nv and "bruto" in nv:
                col_map["valor_negociado_bruto"] = i
            elif "negociado" in nv and "unitario" not in nv and "bruto" not in nv and "liquido" not in nv:
                col_map["valor_negociado"] = i
            else:
                # Tenta detectar coluna de mês (FEV, MAR, ABR...)
                m = _month_from_str(nv)
                if m is not None:
                    month_cols[m] = i

        if "plataforma" not in col_map:
            continue

        # Adiciona colunas de mês ao col_map
        for m, c in month_cols.items():
            col_map[f"month_{m}"] = c

        # Detecta o formato
        has_months = bool(month_cols)
        has_periodo = "periodo" in col_map
        has_insercoes = "insercoes" in col_map

        if has_months and ("marca" in col_map or "projeto" in col_map):
            fmt = "globo"
        elif has_periodo and has_insercoes:
            fmt = "valor"
        elif has_insercoes and "titulo" in col_map:
            fmt = "valor"
        else:
            continue  # não reconhece o formato

        return r, fmt, col_map

    return None


# ─── Parser formato Globo (colunas mensais) ───────────────────────────────────

def _parse_globo_format(
    ws: Any,
    header_row: int,
    col_map: dict[str, int],
    year: int,
    sheet_name: str,
) -> list[dict[str, Any]]:
    """
    Cada linha = uma entrega. Colunas FEV/MAR/ABR... indicam inserções por mês.
    VALOR TOTAL NEGOCIADO = custo total da entrega (dividido proporcionalmente entre os meses).
    """
    results: list[dict[str, Any]] = []

    month_cols = {
        int(k.split("_")[1]): v
        for k, v in col_map.items()
        if k.startswith("month_")
    }

    # Coluna de custo: preferência → negociado > negociado_bruto > tabela
    cost_col = (
        col_map.get("valor_negociado")
        or col_map.get("valor_negociado_bruto")
        or col_map.get("valor_tabela")
    )

    last_seen: dict[str, str] = {}

    for r in range(header_row + 1, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]

        def get(key: str) -> Any:
            col = col_map.get(key)
            if col is None or col - 1 >= len(row):
                return None
            return row[col - 1]

        def get_str(key: str, carry: bool = True) -> str:
            """Lê célula como string; se vazia e carry=True, reutiliza valor anterior."""
            raw = str(get(key) or "").strip()
            if raw:
                last_seen[key] = raw
                return raw
            return last_seen.get(key, "") if carry else ""

        plataforma = get_str("plataforma")
        formato = get_str("formato", carry=False)
        marca = get_str("marca")
        caderno = get_str("caderno")
        projeto = get_str("projeto")

        # Pula linhas vazias ou de total
        if not plataforma and not formato:
            continue
        n_fmt = _norm(formato)
        n_plat = _norm(plataforma)
        if any(k in n_fmt for k in ("total", "subtotal")) or any(k in n_plat for k in ("total", "subtotal")):
            continue

        # Meses com inserções
        month_insertions: list[tuple[int, int]] = []
        for m, c in sorted(month_cols.items()):
            if c - 1 < len(row):
                ins = _parse_int(row[c - 1])
                if ins and ins > 0:
                    month_insertions.append((m, ins))

        if not month_insertions:
            continue

        # Custo total da entrega
        total_cost = _parse_cost(row[cost_col - 1] if cost_col and cost_col - 1 < len(row) else None)

        # Distribuição de custo proporcional às inserções
        total_ins = sum(ins for _, ins in month_insertions)
        day_costs: list[float | None] = []
        day_impressions: list[int | None] = []

        for m, ins in month_insertions:
            if total_cost is not None and total_ins > 0:
                day_costs.append(round(total_cost * ins / total_ins, 2))
            else:
                day_costs.append(None)
            day_impressions.append(None)

        days = [[date(year, m, 1).isoformat(), ins] for m, ins in month_insertions]
        media_type, media_channel = _plataforma_to_channel(plataforma)

        results.append({
            "sheet": sheet_name,
            "media_type": media_type,
            "media_channel": media_channel,
            "data": {
                "market": marca or sheet_name,
                "channel": plataforma[:100],
                "program": projeto[:150],
                "property_text": caderno[:250],
                "format_text": formato[:250],
                "duration_sec": None,
                "external_ref": "",
                "start_date": None,
                "end_date": None,
                "valor_negociado": total_cost,
                "day_costs": day_costs,
                "day_impressions": day_impressions,
            },
            "days": days,
            "piece_codes": [],
        })

    return results


# ─── Parser formato Valor (PERIODO + INSERCOES) ───────────────────────────────

def _parse_valor_format(
    ws: Any,
    header_row: int,
    col_map: dict[str, int],
    year: int,
    sheet_name: str,
) -> list[dict[str, Any]]:
    """
    Cada linha = uma entrega com período (número do mês) e quantidade de inserções.
    VALOR TOTAL NEGOCIADO BRUTO/LIQUIDO = custo da entrega.
    MEDIA DE IMPACTOS = impressões estimadas.
    """
    results: list[dict[str, Any]] = []

    # Coluna de custo
    cost_col = (
        col_map.get("valor_negociado_bruto")
        or col_map.get("valor_negociado")
        or col_map.get("valor_tabela")
    )
    impressions_col = col_map.get("media_impactos")

    for r in range(header_row + 1, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]

        def get(key: str) -> Any:
            col = col_map.get(key)
            if col is None or col - 1 >= len(row):
                return None
            return row[col - 1]

        plataforma = str(get("plataforma") or "").strip()
        formato = str(get("formato") or "").strip()
        acao = str(get("acao") or get("titulo") or "").strip()
        periodo_raw = get("periodo")
        insercoes_raw = get("insercoes")

        # Pula linhas vazias ou de total
        if not plataforma and not formato:
            continue
        if any(k in _norm(formato) for k in ("total", "subtotal")):
            continue
        if any(k in _norm(acao) for k in ("total", "subtotal")):
            continue

        # Mês do período
        month = _parse_int(periodo_raw)
        if month is None:
            month = _month_from_str(periodo_raw)
        if month is None or not (1 <= month <= 12):
            continue

        insertions = _parse_int(insercoes_raw) or 0
        if insertions <= 0:
            continue

        # Custo
        cost_val = row[cost_col - 1] if cost_col and cost_col - 1 < len(row) else None
        total_cost = _parse_cost(cost_val)

        # Impressões estimadas
        imp_val = row[impressions_col - 1] if impressions_col and impressions_col - 1 < len(row) else None
        impressions = _parse_int(imp_val)

        media_type, media_channel = _plataforma_to_channel(plataforma)

        day_iso = date(year, month, 1).isoformat()

        results.append({
            "sheet": sheet_name,
            "media_type": media_type,
            "media_channel": media_channel,
            "data": {
                "market": sheet_name,
                "channel": plataforma[:100],
                "program": acao[:150],
                "property_text": "",
                "format_text": formato[:250],
                "duration_sec": None,
                "external_ref": "",
                "start_date": None,
                "end_date": None,
                "valor_negociado": total_cost,
                "day_costs": [total_cost],
                "day_impressions": [impressions],
            },
            "days": [[day_iso, insertions]],
            "piece_codes": [],
        })

    return results


# ─── Entrada principal ────────────────────────────────────────────────────────

def main() -> int:
    if len(sys.argv) < 2:
        return 2

    path = sys.argv[1]

    try:
        import importlib.abc
        import importlib.machinery

        class _BlockLxml(importlib.abc.MetaPathFinder, importlib.abc.Loader):
            def find_spec(self, fullname, path, target=None):
                if fullname == "lxml" or fullname.startswith("lxml."):
                    return importlib.machinery.ModuleSpec(fullname, self)
                return None
            def create_module(self, spec):
                return None
            def exec_module(self, module):
                raise ImportError("lxml disabled")

        sys.meta_path.insert(0, _BlockLxml())
    except Exception:
        pass

    import openpyxl  # type: ignore

    wb = openpyxl.load_workbook(path, data_only=True)
    parsed_rows: list[dict[str, Any]] = []
    total_rows = 0
    detected: dict[str, Any] = {"sheets": {}}

    for ws in wb.worksheets:
        if _should_skip_sheet(ws.title):
            detected["sheets"][ws.title] = {"skipped": True, "reason": "sheet de resumo/vazio"}
            continue

        year = _infer_year(ws, until_row=10)
        result = _find_sponsorship_header(ws)

        if result is None:
            detected["sheets"][ws.title] = {
                "error": "Cabeçalho de patrocínio não encontrado (sem coluna PLATAFORMA)."
            }
            continue

        header_row, fmt, col_map = result
        detected["sheets"][ws.title] = {
            "format": fmt,
            "header_row": header_row,
            "year": year,
            "columns": {k: v for k, v in col_map.items()},
        }

        if fmt == "globo":
            rows = _parse_globo_format(ws, header_row, col_map, year, ws.title)
        else:
            rows = _parse_valor_format(ws, header_row, col_map, year, ws.title)

        total_rows += len(rows)
        parsed_rows.extend(rows)

    out = {
        "sheets": [ws.title for ws in wb.worksheets],
        "total_rows": total_rows,
        "detected": detected,
        "rows": parsed_rows,
        "pieces": [],
        "format": "sponsorship",
    }
    sys.stdout.write(json.dumps(out, ensure_ascii=False, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
