from __future__ import annotations

from datetime import date, datetime, time
import json
import re
import sys
from typing import Any


class DateTimeEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.isoformat()
        if isinstance(obj, date):
            return obj.isoformat()
        if isinstance(obj, time):
            return obj.isoformat()
        return super().default(obj)


_MONTHS: dict[str, int] = {
    "janeiro": 1,
    "jan": 1,
    "fevereiro": 2,
    "fev": 2,
    "march": 3,
    "marco": 3,
    "março": 3,
    "mar": 3,
    "abril": 4,
    "abr": 4,
    "maio": 5,
    "mai": 5,
    "june": 6,
    "junho": 6,
    "jun": 6,
    "july": 7,
    "julho": 7,
    "jul": 7,
    "agosto": 8,
    "ago": 8,
    "september": 9,
    "setembro": 9,
    "set": 9,
    "october": 10,
    "outubro": 10,
    "out": 10,
    "november": 11,
    "novembro": 11,
    "nov": 11,
    "december": 12,
    "dezembro": 12,
    "dez": 12,
}


def _disable_lxml() -> None:
    try:
        import importlib.abc
        import importlib.machinery
    except Exception:
        return

    class _BlockLxml(importlib.abc.MetaPathFinder, importlib.abc.Loader):
        def find_spec(self, fullname: str, path: Any, target: Any = None) -> Any:
            if fullname == "lxml" or fullname.startswith("lxml."):
                return importlib.machinery.ModuleSpec(fullname, self)
            return None

        def create_module(self, spec: Any) -> Any:
            return None

        def exec_module(self, module: Any) -> None:
            raise ImportError("lxml disabled")

    sys.meta_path.insert(0, _BlockLxml())


def _norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = (
        s.replace("á", "a")
        .replace("à", "a")
        .replace("â", "a")
        .replace("ã", "a")
        .replace("é", "e")
        .replace("ê", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ô", "o")
        .replace("õ", "o")
        .replace("ú", "u")
        .replace("ç", "c")
    )
    return s


def detect_media_from_sheet(sheet_name: str) -> tuple[str, str]:
    n = _norm(sheet_name)
    if "open tv" in n or "tv aberta" in n:
        return ("offline", "tv_aberta")
    if "paytv" in n or "pay tv" in n:
        return ("offline", "paytv")
    if "radio" in n:
        return ("offline", "radio")
    if "jornal" in n:
        return ("offline", "jornal")
    if "ooh" in n:
        return ("offline", "ooh")
    if "meta" in n:
        return ("online", "meta")
    if "google" in n:
        return ("online", "google")
    if "youtube" in n:
        return ("online", "youtube")
    if "display" in n:
        return ("online", "display")
    if "search" in n:
        return ("online", "search")
    if "social" in n or "digital" in n:
        return ("online", "social")
    return ("online", "other")


def _try_parse_date(v: Any) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass
    return None


def _try_parse_datetime(v: Any) -> datetime | None:
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt)
            except Exception:
                pass
    return None


def _parse_int(v: Any) -> int | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, int):
        return int(v)
    if isinstance(v, float):
        if v.is_integer():
            return int(v)
        return int(round(v))
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        s = s.replace(",", ".")
        try:
            f = float(s)
            return int(round(f))
        except Exception:
            return None
    return None


def _maybe_month(v: Any) -> int | None:
    if v is None:
        return None
    s = _norm(str(v))
    if not s:
        return None
    for name, num in _MONTHS.items():
        if name in s:
            return num
    return None


def _infer_year(ws: Any, *, until_row: int) -> int:
    for r in range(1, max(1, min(until_row, ws.max_row)) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v)
            m4 = re.search(r"\b(20\d{2})\b", s)
            if m4:
                y = int(m4.group(1))
                if 2000 <= y <= 2100:
                    return y
            m2 = re.search(r"\b(\d{2})\s*/\s*(\d{2})\b", s)
            if m2:
                y2 = int(m2.group(2))
                if 0 <= y2 <= 99:
                    return 2000 + y2
    return datetime.now().year


def _find_table_header_row(ws: Any) -> int | None:
    for r in range(1, min(ws.max_row, 200) + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        tokens = [_norm(str(v)) for v in row_vals if v is not None and str(v).strip() != ""]
        if not tokens:
            continue
        if "market" in tokens and "channel" in tokens:
            return r
    return None


def _find_day_row(ws: Any, *, start_row: int) -> int | None:
    best_row = None
    best_count = 0
    for r in range(start_row, min(ws.max_row, start_row + 12) + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        count = 0
        for v in row_vals:
            n = _parse_int(v)
            if n is not None and 1 <= n <= 31:
                count += 1
        if count > best_count:
            best_count = count
            best_row = r
    if best_row is None or best_count < 7:
        return None
    return best_row


def _extract_piece_table(ws: Any) -> list[dict[str, Any]]:
    header_row = None
    code_col = None
    title_col = None
    sec_col = None

    for r in range(1, min(ws.max_row, 120) + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 200) + 1)]
        norm_vals = [_norm(str(v)) if v is not None else "" for v in row_vals]
        if "pc" not in norm_vals:
            continue
        if "titulos" not in norm_vals and "titulos" not in " ".join(norm_vals):
            continue
        if "sec" not in norm_vals:
            continue
        header_row = r
        for c, nv in enumerate(norm_vals, start=1):
            if nv == "pc":
                code_col = c
            elif nv == "sec":
                sec_col = c
            elif "titulo" in nv:
                title_col = c
        if title_col is None:
            for c, nv in enumerate(norm_vals, start=1):
                if "titulo" in nv:
                    title_col = c
                    break
        break

    if header_row is None or code_col is None or title_col is None:
        return []

    pieces: list[dict[str, Any]] = []
    for r in range(header_row + 1, min(ws.max_row, header_row + 80) + 1):
        code = ws.cell(row=r, column=code_col).value
        title = ws.cell(row=r, column=title_col).value
        sec = ws.cell(row=r, column=sec_col).value if sec_col is not None else None
        code_s = str(code).strip().upper() if code is not None else ""
        title_s = str(title).strip() if title is not None else ""
        if not code_s and not title_s:
            break
        if not code_s:
            continue
        dur = _parse_int(sec) if sec is not None else None
        pieces.append({"code": code_s, "title": title_s, "duration_sec": dur})
    return pieces


def _split_piece_codes(v: Any) -> list[str]:
    if v is None:
        return []
    s = str(v).strip()
    if not s:
        return []
    tokens = re.split(r"[^A-Za-z0-9]+", s)
    codes: list[str] = []
    for t in tokens:
        t = t.strip().upper()
        if not t or len(t) > 12:
            continue
        codes.append(t)
    return sorted(set(codes))


def main() -> int:
    if len(sys.argv) < 2:
        return 2
    path = sys.argv[1]
    _disable_lxml()
    import openpyxl  # type: ignore

    wb = openpyxl.load_workbook(path, data_only=True)
    parsed_rows: list[dict[str, Any]] = []
    detected: dict[str, Any] = {"sheets": {}}
    total_rows = 0
    pieces_map: dict[str, dict[str, Any]] = {}

    for ws in wb.worksheets:
        for p in _extract_piece_table(ws):
            code = str(p.get("code") or "").strip().upper()
            if not code:
                continue
            if code not in pieces_map:
                pieces_map[code] = p
            else:
                if p.get("title") and not pieces_map[code].get("title"):
                    pieces_map[code]["title"] = p.get("title")
                if p.get("duration_sec") and not pieces_map[code].get("duration_sec"):
                    pieces_map[code]["duration_sec"] = p.get("duration_sec")
        media_type, media_channel = detect_media_from_sheet(ws.title)
        header_row_idx = _find_table_header_row(ws)
        if header_row_idx is None:
            detected["sheets"][ws.title] = {
                "media_type": media_type,
                "media_channel": media_channel,
                "header_row": None,
                "day_row": None,
                "columns": {},
                "date_columns_count": 0,
                "error": "Tabela não encontrada (não achei cabeçalhos MARKET/CHANNEL).",
            }
            continue

        header_values = [ws.cell(row=header_row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        day_row_idx = _find_day_row(ws, start_row=header_row_idx)
        year_hint = _infer_year(ws, until_row=header_row_idx)

        headers = [_norm(str(v)) if v is not None else "" for v in header_values]
        col_by_key: dict[str, int] = {}
        date_cols: list[tuple[int, date]] = []
        for idx, h in enumerate(headers, start=1):
            if not h:
                continue
            if h in {"praca", "praça", "market"} or "praca" in h or "praça" in h or h == "mercado":
                col_by_key["market"] = idx
            elif h in {"canal", "channel"}:
                col_by_key["channel"] = idx
            elif h in {"programa", "program"}:
                col_by_key["program"] = idx
            elif h in {"janela", "property", "segmentacao", "segmentacao/posicionamento"} or "janela" in h:
                col_by_key["property_text"] = idx
            elif h in {"formato", "format"}:
                col_by_key["format_text"] = idx
            elif h in {"duracao", "duracao(s)", "duration", "duration_sec"} or "duracao" in h:
                col_by_key["duration_sec"] = idx
            elif h in {"external_ref", "external ref", "id", "chave", "linha_id"} or "external" in h:
                col_by_key["external_ref"] = idx
            elif h in {"inicio", "início", "start", "start_date", "data_inicio"} or "inicio" in h:
                col_by_key["start_date"] = idx
            elif h in {"fim", "end", "end_date", "data_fim"}:
                col_by_key["end_date"] = idx
            elif h in {"peca", "peça", "criativo", "creative"}:
                col_by_key["piece_codes"] = idx
            else:
                maybe_date = _try_parse_date(header_values[idx - 1])
                if maybe_date is not None:
                    date_cols.append((idx, maybe_date))

        if day_row_idx is not None:
            day_values = [ws.cell(row=day_row_idx, column=c).value for c in range(1, ws.max_column + 1)]
            month_by_col: dict[int, int] = {}
            for col_idx in range(1, ws.max_column + 1):
                month = None
                for r in range(header_row_idx, day_row_idx):
                    last = None
                    for c in range(1, col_idx + 1):
                        mv = _maybe_month(ws.cell(row=r, column=c).value)
                        if mv is not None:
                            last = mv
                    if last is not None:
                        month = last
                if month is not None:
                    month_by_col[col_idx] = month

            # Ajustar anos para transições (ex: Dez 2025 -> Jan 2026)
            # Se year_hint=2026 e temos dezembro antes de janeiro, dezembro é 2025
            year_by_month: dict[int, int] = {}
            sorted_cols = sorted(month_by_col.keys())
            if sorted_cols:
                # Detectar se há transição de ano (mês decresce na sequência)
                months_in_order = [month_by_col[c] for c in sorted_cols]
                unique_months = []
                for m in months_in_order:
                    if not unique_months or unique_months[-1] != m:
                        unique_months.append(m)

                # Se sequência é tipo [12, 1, 2], dezembro é ano anterior
                # Se sequência é tipo [1, 2, 3], todos são mesmo ano
                current_year = year_hint
                for i, m in enumerate(unique_months):
                    if i > 0 and m < unique_months[i - 1]:
                        # Virada de ano detectada, meses anteriores são ano anterior
                        for prev_m in unique_months[:i]:
                            year_by_month[prev_m] = year_hint - 1
                        current_year = year_hint
                    if m not in year_by_month:
                        year_by_month[m] = current_year

            for col_idx, v in enumerate(day_values, start=1):
                n = _parse_int(v)
                if n is None or not (1 <= n <= 31):
                    continue
                month = month_by_col.get(col_idx) or datetime.now().month
                year = year_by_month.get(month, year_hint)
                try:
                    date_cols.append((col_idx, date(year, month, n)))
                except Exception:
                    pass

        detected["sheets"][ws.title] = {
            "media_type": media_type,
            "media_channel": media_channel,
            "header_row": header_row_idx,
            "day_row": day_row_idx,
            "year_hint": year_hint,
            "columns": col_by_key,
            "date_columns_count": len(date_cols),
        }

        start_data_row = (day_row_idx + 1) if day_row_idx is not None else (header_row_idx + 1)
        last_seen: dict[str, str] = {}
        for r in range(start_data_row, ws.max_row + 1):
            total_rows += 1
            row_values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            if not any(v is not None and str(v).strip() != "" for v in row_values):
                continue

            data: dict[str, Any] = {}
            data["market"] = (row_values[col_by_key["market"] - 1] if "market" in col_by_key else "") or ""
            data["channel"] = (row_values[col_by_key["channel"] - 1] if "channel" in col_by_key else "") or ""
            data["program"] = (row_values[col_by_key["program"] - 1] if "program" in col_by_key else "") or ""
            data["property_text"] = (
                (row_values[col_by_key["property_text"] - 1] if "property_text" in col_by_key else "") or ""
            )
            data["format_text"] = (row_values[col_by_key["format_text"] - 1] if "format_text" in col_by_key else "") or ""

            # Verificar se a linha é um cabeçalho de seção (ex: "TT PAY TV", "OPEN TV")
            # Esses valores não devem ser tratados como market/cidade
            market_val = str(data.get("market") or "").strip()
            market_norm = _norm(market_val)
            is_section_header = any(
                pattern in market_norm
                for pattern in ("pay tv", "paytv", "open tv", "tv aberta", "radio", "jornal", "ooh", "digital", "meta", "google", "youtube")
            )
            if is_section_header:
                data["market"] = ""  # Não usar como market

            for k in ("market", "channel", "program", "property_text", "format_text"):
                v = str(data.get(k) or "").strip()
                if not v and k in last_seen:
                    data[k] = last_seen[k]
                elif v:
                    last_seen[k] = v

            duration_raw = row_values[col_by_key["duration_sec"] - 1] if "duration_sec" in col_by_key else None
            data["duration_sec"] = _parse_int(duration_raw)
            data["external_ref"] = (
                (row_values[col_by_key["external_ref"] - 1] if "external_ref" in col_by_key else "") or ""
            )
            start_raw = row_values[col_by_key["start_date"] - 1] if "start_date" in col_by_key else None
            end_raw = row_values[col_by_key["end_date"] - 1] if "end_date" in col_by_key else None
            start_dt = _try_parse_datetime(start_raw)
            end_dt = _try_parse_datetime(end_raw)
            data["start_date"] = start_dt.isoformat() if start_dt else None
            data["end_date"] = end_dt.isoformat() if end_dt else None

            piece_codes: list[str] = []
            if "piece_codes" in col_by_key:
                piece_codes = _split_piece_codes(row_values[col_by_key["piece_codes"] - 1])

            days: list[list[Any]] = []
            for col_idx, d in date_cols:
                v = row_values[col_idx - 1] if col_idx - 1 < len(row_values) else None
                if v is None:
                    continue

                # Verificar se é um código de peça (letra) ou número de inserções
                v_str = str(v).strip().upper()
                if v_str and len(v_str) <= 2 and v_str.isalpha():
                    # É um código de peça (ex: A, C, D, E, F, G, H)
                    if v_str not in piece_codes:
                        piece_codes.append(v_str)
                    days.append([d.isoformat(), 1])  # 1 inserção para esta peça/data
                else:
                    # É um número de inserções
                    ins = _parse_int(v)
                    if ins is None or ins <= 0:
                        continue
                    days.append([d.isoformat(), ins])

            # Pular linhas sem dados úteis
            if (
                not str(data.get("market") or "").strip()
                and not str(data.get("channel") or "").strip()
                and not str(data.get("program") or "").strip()
                and not str(data.get("property_text") or "").strip()
                and not str(data.get("format_text") or "").strip()
                and not str(data.get("external_ref") or "").strip()
            ):
                continue

            # Pular linhas que são apenas cabeçalhos de seção (sem canal/programa/dias)
            if is_section_header and not days and not str(data.get("channel") or "").strip() and not str(data.get("program") or "").strip():
                continue

            parsed_rows.append(
                {
                    "sheet": ws.title,
                    "media_type": media_type,
                    "media_channel": media_channel,
                    "data": data,
                    "days": days,
                    "piece_codes": piece_codes,
                }
            )

    out = {
        "sheets": [ws.title for ws in wb.worksheets],
        "total_rows": total_rows,
        "detected": detected,
        "rows": parsed_rows,
        "pieces": list(pieces_map.values()),
    }
    sys.stdout.write(json.dumps(out, ensure_ascii=False, cls=DateTimeEncoder))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
